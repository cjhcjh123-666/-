import sys
from time import sleep
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
import pyautogui
import pyperclip
import json
from selenium.webdriver import Chrome
from selenium.webdriver import ChromeOptions
import datetime
import schedule
from chaojiying import Chaojiying_Client
option = ChromeOptions()
option.add_experimental_option('excludeSwitches', ['enable-automation'])
# option.add_argument('"--proxy-server=http://'+ proxy)
driver = Chrome(options=option)
#控制浏览器访问url地址
driver.get("https://yuyues.hnmuseum.com/w/home?stack-key=24eba665")
driver.maximize_window()
def getcookie():
    # 程序打开网页后20秒内手动登陆账户
    sleep(20)

    with open('cookies.txt', 'w') as cookief:
        # 将cookies保存为json格式
        cookief.write(json.dumps(driver.get_cookies()))
def login():
    with open('cookies.txt', 'r', encoding='utf8') as f:
        listCookies = json.loads(f.read())

    # 往browser里添加cookies
    for cookie in listCookies:
        cookie_dict = {
            'domain': 'yuyues.hnmuseum.com',
            'name': cookie.get('name'),
            'value': cookie.get('value'),
            "sameSite": cookie.get('sameSite'),
            'path': '/',
            'httpOnly': False,
            #'HostOnly': False,
            'Secure': False
        }
        driver.add_cookie(cookie_dict)
    driver.refresh()  # 刷新网页,cookies才成功
def add():
    sleep(0.5)
    try:
        a=driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[1]/div[1]/div[4]/div/div/a[2]')
        a.click()
        print('你已经登录')
    except Exception as e:
        print('点击个人中心失败，请坚持是否登录')
    sleep(0.5)
    try:
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/ul/li[4]/a')
        a.click()
    except Exception as e:
        print('找不到常用联系人元素')
    workbook=openpyxl.load_workbook("sfz.xlsx")
    sfz=workbook['Sheet1']
    for i in range(1,6):
        try:
            a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[1]/a')
            a.click()
        except Exception as e:
            print('找不到add元素')
        sleep(1)

        try:
            location = pyautogui.locateCenterOnScreen("name.png", confidence=0.7)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button='left')
            pyperclip.copy(sfz.cell(row=i, column=1).value)
            pyautogui.hotkey('ctrl', 'v')
        except Exception as e:
            print('找不到输入框1元素'+str(e))
        try:
            location = pyautogui.locateCenterOnScreen("haoma.png", confidence=0.7)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button='left')
            pyperclip.copy(sfz.cell(row=i, column=2).value)
            pyautogui.hotkey('ctrl', 'v')
        except Exception as e:
            print('找不到输入框2元素' + str(e))

        try:
            #driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[3]/div/div/div/div[2]/a').click()
            a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[3]/div/div/div/div[2]/a')
            a.click()
        except Exception as e:
            print('找不到确认元素'+str(e))
getcookie()
login()
add()