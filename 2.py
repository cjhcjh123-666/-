import sys
from time import sleep
import locale
from selenium import webdriver
import openpyxl
from selenium.webdriver.common.by import By
import pyautogui
import pyperclip
import json
from selenium.webdriver import Chrome
from selenium.webdriver import ChromeOptions
import datetime
import schedule
from chaojiying import Chaojiying_Client
print('初次使用，请先登录,请选择1.只新增联系人  3.爷要抢票（有常用联系人）')
s=input()
if s=='3' or  s=='test':
    ss=int(input('几位爷啊？'))
# proxy='202.20.16.82:10152'
option = ChromeOptions()
option.add_experimental_option('excludeSwitches', ['enable-automation'])
# option.add_argument('"--proxy-server=http://'+ proxy)
driver = Chrome(options=option)
driver.implicitly_wait(10)
#控制浏览器访问url地址
driver.get("https://yuyues.hnmuseum.com/w/home?stack-key=24eba665")
driver.maximize_window()
cg=0
def getcookie():
    # 程序打开网页后20秒内手动登陆账户
    time.sleep(20)

    with open('cookies.txt', 'w') as cookief:
        # 将cookies保存为json格式
        cookief.write(json.dumps(driver.get_cookies()))
def login():
    with open('cookies.txt', 'r', encoding='utf8') as f:
        listCookies = json.loads(f.read())

    # 往browser里添加cookies
    for cookie in listCookies:
        cookie_dict = {
            'domain': 'yuyues.hnmuseum.com',
            'name': cookie.get('name'),
            'value': cookie.get('value'),
            "sameSite": cookie.get('sameSite'),
            'path': '/',
            'httpOnly': False,
            #'HostOnly': False,
            'Secure': False
        }
        driver.add_cookie(cookie_dict)
    driver.refresh()  # 刷新网页,cookies才成功

    #
    # try:
    #     driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[1]/div[1]/div[4]/div/div/a').click()
    #     print('请扫码登录微信')
    #     sleep(10)
    # except Exception as e:
    #     print('登录失败啦')
def add():
    sleep(0.5)
    try:
        a=driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[1]/div[1]/div[4]/div/div/a[2]')
        a.click()
        print('你已经登录')
    except Exception as e:
        print('点击个人中心失败，请坚持是否登录')
    sleep(0.5)
    try:
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/ul/li[4]/a')
        a.click()
    except Exception as e:
        print('找不到常用联系人元素')
    workbook=openpyxl.load_workbook("sfz.xlsx")
    sfz=workbook['Sheet1']

    for i in range(1,6):
        try:
            a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[1]/a')
            a.click()
        except Exception as e:
            print('找不到add元素')
        sleep(1)

        try:
            location = pyautogui.locateCenterOnScreen("name.png", confidence=0.7)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button='left')
            pyperclip.copy(sfz.cell(row=i, column=1).value)
            pyautogui.hotkey('ctrl', 'v')
        except Exception as e:
            print('找不到输入框1元素'+str(e))
        try:
            location = pyautogui.locateCenterOnScreen("haoma.png", confidence=0.7)
            if location is not None:
                pyautogui.click(location.x, location.y, clicks=1, interval=0.2, duration=0.2, button='left')
            pyperclip.copy(sfz.cell(row=i, column=2).value)
            pyautogui.hotkey('ctrl', 'v')
        except Exception as e:
            print('找不到输入框2元素' + str(e))

        try:
            #driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[3]/div/div/div/div[2]/a').click()
            a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[3]/div/div/div/div[2]/a')
            a.click()
        except Exception as e:
            print('找不到确认元素'+str(e))
# def delete(n):
#     sleep(0.5)
#     try:
#         driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[1]/div[1]/div[4]/div/div/a[2]').click()
#         print('你已经登录')
#     except Exception as e:
#         print('点击个人中心失败，请坚持是否登录')
#     sleep(0.5)
#     try:
#         driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/ul/li[4]/a').click()
#     except Exception as e:
#         print('找不到常用联系人元素')
#     sleep(0.5)
#     #print(n,type(n))
#     for i in range(n):
#         try:
#             button=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/div/div[2]/table/tbody/tr[1]/td[4]/a[2]')
#             driver.execute_script("(arguments[0]).click();", button)
#         except Exception as e:
#             print(e)
def getdate(b):
    locale.setlocale(locale.LC_CTYPE, 'Chinese')
    a = datetime.date.strftime(b, '%m月%d日')
    ri = ''
    yue = ''
    for i in range(len(a)):
        if a[i] != '月':
            yue += a[i]
        else:
            for j in range(i + 1, len(a)):
                if a[j] != '日':
                    ri += a[j]
            break

    return str(str(int(yue)) + "月" + str(int(ri)) + '日')
def booktest(ss):
    global cg
    try:
        workbook=openpyxl.load_workbook("qp.xlsx")
        qp=workbook['Sheet1']
        n=qp.max_row
        for i in range(1,n):
            sleep(0.5)
            #global f, aim
            aim = 0
            try:
                a=driver.find_element(by=By.XPATH, value='//*[@id="app"]/div/div[1]/div[1]/div[3]/a[1]')
                a.click()
                print('点击参观预约')
            except Exception as e:
                print('找不到参观预约元素')

            try:
                a=driver.find_element(by=By.XPATH,
                                    value='//*[@id="app"]/div/div[2]/div[3]/div[1]/div[2]/a[1]/span')
                a.click()
                print('点击个人预约')
            except Exception as e:
                print('找不到个人预约元素')
            sleep(2)
            if qp.cell(row=i+1,column=4).value=='yes':
                continue
            aim=0
            for j in range(1,15):
                try:
                    time=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[1]/ul[2]/li[{}]/span[1]'.format(j)).text

                    print(time)
                except Exception as e:
                    continue
                if time==getdate(qp.cell(row=i+1,column=1).value):
                    aim=j
                    break
            #print(type(aim),aim)
            if aim!=0:
                for kk in range(10):
                    if driver.find_element(by=By.XPATH,
                                           value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[1]/ul[2]/li[{}]/span[2]'.format(
                                                   aim)).text == '无票':

                        pyautogui.hotkey('f5')
                        sleep(5)
                    else:
                        break
                if driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[1]/ul[2]/li[{}]/span[2]'.format(aim)).text=='无票':
                    continue
                else:
                    driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[1]/ul[2]/li[{}]/span[2]'.format(aim)).click()
                    if qp.cell(row=i+1,column=2).value=='上午':
                        if driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[1]').get_attribute('class')\
                                =='open'and((driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[1]/p[2]/span[1]'
                                                                 ).get_attribute('class')!='no'
                                             and qp.cell(row=i+1,column=3).value=='基础') or
                                            (driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[1]/p[2]/span[2]'
                                                                 ).get_attribute('class')!='no' and qp.cell(row=i+1,column=3).value=='临时')):
                            try:
                                driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[1]').click()
                                print('上午')
                            except Exception as e:
                                print('莫名其妙出错')
                            sleep(0.5)
                            driver.find_element(by=By.XPATH,
                                                value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/p/span[1]/i').click()
                            driver.find_element(by=By.XPATH,
                                                value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/a').click()
                            sleep(0.5)
                            addguanzhong(i,qp,ss)
                            sleep(0.3)
                            submit()
                            if qp.cell(row=i+1,column=3).value=='临时':
                                driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]').click()
                            qp.cell(row=i + 1, column=4, value='yes')
                            cg+=1
                            workbook.save('qp.xlsx')
                        else:
                            continue

                    elif qp.cell(row=i+1,column=2).value=='下午':
                        print(driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]'
                                               ).get_attribute('class')=='open')
                        print((driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]/p[2]/span[1]'
                                                                 ).get_attribute('class')!='no'))
                        print(qp.cell(row=i+1,column=3).value=='基础')
                        print((driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]/p[2]/span[2]'
                                                                 ).get_attribute('class')!='no'))
                        print(qp.cell(row=i+1,column=3).value=='临时')
                        if driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]'
                                               ).get_attribute('class')=='open'and(
                                (driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]/p[2]/span[1]'
                                                                 ).get_attribute('class')!='no'
                                             and qp.cell(row=i+1,column=3).value=='基础') or
                                            (driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]/p[2]/span[2]'
                                                                 ).get_attribute('class')!='no' and qp.cell(row=i+1,column=3).value=='临时')):
                            try:
                                a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[1]/div[2]/ul/li[2]')
                                a.click()
                                print('下午')
                            except Exception as e:
                                print('下午出错')
                            sleep(0.5)
                            a=driver.find_element(by=By.XPATH,
                                                value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/p/span[1]/i')
                            a.click()
                            a=driver.find_element(by=By.XPATH,
                                                value='//*[@id="app"]/div/div[2]/div/div[2]/div[2]/a')
                            a.click()
                            sleep(0.5)
                            addguanzhong(i, qp,ss)
                            sleep(0.3)
                            submit()
                            sleep(0.5)
                            if qp.cell(row=i+1,column=3).value=='临时':
                                a=driver.find_element(by=By.XPATH,value='/html/body/div[1]/div/div[2]/div/div[2]/div[5]/a[2]')
                                a.click()
                            qp.cell(row=i+1,column=4,value='yes')
                            cg += 1
                            workbook.save('qp.xlsx')
                        else:
                            continue
        driver.quit()
    except Exception as e:
        print('找不到元素'+e)
def addguanzhong(k,qp,ss):
    global cg
    workbook = openpyxl.load_workbook("sfz.xlsx")
    sfz = workbook['Sheet1']
    n=sfz.max_row
    for i in range(ss-1):
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a')
        a.click()
    for i in range(cg+1,ss-cg+1):
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[3]/div/table[{}]/tbody/tr/td[2]/a'.format(i))
        a.click()
        sleep(0.5)
        button=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[9]/div/div/div/div[2]/div/ul/li[{}]/div[1]/a[1]'.format(i))
        driver.execute_script("(arguments[0]).click();", button)
        sleep(0.5)
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[9]/div/div/div/div[2]/div/div/a[2]')
        a.click()
    for i in range(1,ss+1):
        if qp.cell(row=k+1,column=3).value=='基础':
            a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[3]/div/table[{}]/tbody/tr/td[5]/ul/li[1]/a/i'.format(i))
            a.click()
        elif qp.cell(row=k+1,column=3).value=='临时':
            a=driver.find_element(by=By.XPATH,
                                value='//*[@id="app"]/div/div[2]/div/div[2]/div[3]/div/table[{}]/tbody/tr/td[5]/ul/li[2]/a/i'.format(
                                    i))
            a.click()
        sleep(0.1)
    a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[8]/p[2]/a[2]')
    a.click()
def submit():
    #（679，383）
    #(1216, 800)
    a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]')
    a.click()
    sleep(0.6)
    from pyautogui import screenshot
    shot = screenshot(region=(679,383, 1216-679, 800-383))
    shot.save('yzm.png')
    # code= open('yzm.png', 'rb').read()
    # ans=chaojiying.PostPic(code,9101)
    # # y=chaojiying.PostPic(code,9101)['pic_str'][1]
    # s=ans['pic_str'].split(',')
    # x=s[0]
    # y=s[1]
    # print(x,y)
    import base64
    import json
    import requests

    # 复制以下代码，只需填入自己的账号密码、待识别的图片路径即可。
    # 关于ID：选做识别的模型ID。

    def b64_api(username, password, img_path, ID):
        with open(img_path, 'rb') as f:
            b64_data = base64.b64encode(f.read())
        b64 = b64_data.decode()
        data = {"username": username, "password": password, "ID": ID, "b64": b64, "version": "3.1.1"}
        data_json = json.dumps(data)
        result = json.loads(requests.post("http://www.tulingtech.xyz/tuling/predict", data=data_json).text)
        return result

    img_path = r"yzm.png"
    result1 = b64_api(username="", password="", img_path=img_path, ID="60392746")
    print(result1)
    x1=result1['data']['顺序1']['X坐标值']
    y1=result1['data']['顺序1']['Y坐标值']
    result2 = b64_api(username="", password="", img_path=img_path, ID="63875840")
    print(result2)
    if result2['code']=='1':
        x2 = result2['data']['顺序1']['X坐标值']
        y2 = result2['data']['顺序1']['Y坐标值']
    try:
        pyautogui.click(x=int(x1+679),y=int(y1+383),button='left',clicks=1,interval=0.0,tween=pyautogui.linear)
    except Exception as e:
        try:
            pyautogui.click(x=int(x2+679), y=int(y2+383), button='left', clicks=1, interval=0.0, tween=pyautogui.linear)
        except Exception as e:
            location1 = pyautogui.locateCenterOnScreen("x.png", confidence=0.7)
            if location1 is not None:
                pyautogui.click(location1.x, location1.y, clicks=1, interval=0.2, duration=0.2, button='left')
            sleep(0.11)
            print(len(driver.find_elements(by=By.XPATH, value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]')))
            if len(driver.find_elements(by=By.XPATH, value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]')) == 1:
                submit()
    sleep(1)
    location1 = pyautogui.locateCenterOnScreen("x.png", confidence=0.7)
    if location1 is not None:
        pyautogui.click(location1.x, location1.y, clicks=1, interval=0.2, duration=0.2, button='left')
    sleep(0.11)
    print(len(driver.find_elements(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]')))
    if len(driver.find_elements(by=By.XPATH,value='//*[@id="app"]/div/div[2]/div/div[2]/div[5]/a[2]'))==1:
        submit()

while True:
    if s=='1':
        login()
        add()
    elif s=='3':
        login()
        booktest(ss)
    elif s=='test':
        login()
        add()
        sleep(0.5)
        a=driver.find_element(by=By.XPATH,value='//*[@id="app"]/div/div[1]/div[1]/div[3]/a[1]')
        a.click()
        booktest(ss)
    elif keyboard.is_pressed('enter'):
        break
'''def job():
    print("I'm working...")

# 每十分钟执行任务
schedule.every(10).minutes.do(job)
# 每个小时执行任务
schedule.every().hour.do(job)
# 每天的10:30执行任务
schedule.every().day.at("10:30").do(job)
# 每个月执行任务
schedule.every().monday.do(job)
# 每个星期三的13:15分执行任务
schedule.every().wednesday.at("13:15").do(job)
# 每分钟的第17秒执行任务
schedule.every().minute.at(":17").do(job)

while True:
    schedule.run_pending()
    time.sleep(1)'''




